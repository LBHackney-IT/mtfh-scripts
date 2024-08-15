from dataclasses import asdict
import json
import uuid
from faker import Faker
import openpyxl.cell
import openpyxl.writer
from aws.database.domain.dynamo_domain_objects import Patch, ResponsibleEntity, ResponsibleEntityContactDetails
from aws.database.dynamodb.utils.get_dynamodb_table import get_dynamodb_table
from aws.utils.logger import Logger
from enums.enums import Stage
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font


class Config:
    STAGE = Stage.HOUSING_DEVELOPMENT
    logger = Logger("tmo_patch_assignment")
    TABLE_NAME = "PatchesAndAreas"
    INPUT_FILE_PATH = f"tmo_patches-{STAGE}.xlsx"
    SHEET_NAME = "TMO managed properties"
    AREA_ID = uuid.uuid4()
    EXPORT_FILE_PATH = f"tmo_patches-{STAGE}.json"


# Create fake data for dev/staging
def create_fake_spreadsheet_data():
    # Generate random data
    print("Creating fake data...")
    fake = Faker()
    df = pd.DataFrame(data=[
        {
            "Block name": fake.address(),
            "TMO": fake.company(),
            "TMO Liaison Officer": (fake.email(), fake.name()),
            "Housing Officer at TMO": (fake.email(), fake.name()),
            "Income Officer": (fake.email(), fake.name()),
            "Housing Services Manager": (fake.email(), fake.name()),
            "Repairs Manager": (fake.email(), fake.name()),
            "TMO Manager": (fake.email(), fake.name())
        } for _ in range(10)
    ])

    # Add a row with the same tmo data but a different housing officer name
    data_row = df.iloc[0].copy()
    data_row["Housing Officer at TMO"] = (fake.email(), fake.name())
    df.loc[len(df.index)] = data_row

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = Config.SHEET_NAME  # type: ignore

    # Add header row to the worksheet
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)  # type: ignore

    # Add data rows to the worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)  # type: ignore
            if isinstance(value, tuple):
                email, name = value
                # Add the hyperlink directly to the cell
                cell.value = name
                cell.hyperlink = f"mailto:{email}"
                cell.font = Font(color="0000FF", underline="single")
            else:
                cell.value = value

    # Save the workbook to an Excel file
    print("Saving fake data to Excel file...")
    wb.save(Config.INPUT_FILE_PATH)


# Extract hyperlinks from an Excel cell
def _get_link_if_exists(cell: openpyxl.cell.Cell) -> tuple[str, str] | tuple[str, None]:
    if cell.hyperlink:
        return (str(cell.value), str(cell.hyperlink.target).removeprefix("mailto:"))
    return (str(cell.value), None)


# Convert the Excel file to a Pandas DataFrame and extract hyperlinks
def extract_hyperlinks_from_xlsx(file_name: str, sheet_name: str, columns_to_parse: list[str], row_header: int = 1) -> pd.DataFrame:
    """
    Extracts hyperlinks from an Excel file and returns a DataFrame.

    Args:
        file_name (str): The path to the Excel file.
        sheet_name (str): The name of the sheet to read from.
        columns_to_parse (list[str]): The list of column names to parse for hyperlinks.
        row_header (int, optional): The row number where the header starts. Defaults to 1.

    Returns:
        pd.DataFrame: The DataFrame with extracted hyperlinks.
    """
    df = pd.read_excel(file_name, sheet_name)
    ws = openpyxl.load_workbook(file_name)[sheet_name]
    for column in columns_to_parse:
        row_offset = row_header + 1
        column_index = list(df.columns).index(column) + 1
        df[column] = [
            _get_link_if_exists(
                ws.cell(row=row_offset + i, column=column_index))
            for i in range(len(df[column]))
        ]
    return df


def build_tmo_area_object() -> Patch:
    return Patch(
        id=str(Config.AREA_ID),
        domain="Housing Management",
        name="TMO Area",
        parentId=str(uuid.UUID(int=0)),  # Empty UUID
        patchType="area",
        responsibleEntities=[],
        versionNumber=0
    )


def build_patch_object(data_row: pd.Series) -> Patch:
    """
    Build a Patch object based on the given data_row.

    Args:
        data_row (pd.Series): The data row containing information for building the Patch object.

    Returns:
        Patch: The built Patch object.
    """
    # Code implementation...
    # Responsible entities
    tmo_managers = data_row["TMO Manager"]  # Set of tuples
    tmo_hos = data_row["Housing Officer at TMO"]  # Set of tuples

    responsible_entities = []
    for tmo_manager in tmo_managers:
        responsible_entities.append(
            ResponsibleEntity(
                id=str(uuid.uuid4()),
                name=tmo_manager[0],
                contactDetails=ResponsibleEntityContactDetails(
                    emailAddress=tmo_manager[1]),
                responsibleType="TmoManager"
            )
        )

    for tmo_ho in tmo_hos:
        responsible_entities.append(
            ResponsibleEntity(
                id=str(uuid.uuid4()),
                name=tmo_ho[0],
                contactDetails=ResponsibleEntityContactDetails(
                    emailAddress=tmo_ho[1]),
                responsibleType="TMO"
            )
        )

    # Create the Patch object
    return Patch(
        id=str(uuid.uuid4()),
        domain="Housing Management",
        name=data_row["TMO"],
        parentId=str(Config.AREA_ID),
        patchType="tmoPatch",
        responsibleEntities=responsible_entities,
        versionNumber=0
    )


def create_tmo_patches():
    """
    Extracts data from an Excel file, creates TMO Patch objects,
    exports the patches to a JSON file, and saves the patches to the database.

    Returns:
        None
    """
    # Get unique patches from the Excel file
    Config.logger.log(
        f"Extracting data from Excel file {Config.INPUT_FILE_PATH}")
    _file_path = Config.INPUT_FILE_PATH
    df = extract_hyperlinks_from_xlsx(_file_path, Config.SHEET_NAME, [
        "Housing Officer at TMO", "TMO Manager"])

    # Group by TMO name and aggregate the responsible entities into sets
    # (Should result in a single row per TMO, with a set of TMOs and a set of TMO Managers)
    df = df.groupby("TMO", as_index=False).agg(
        {"Housing Officer at TMO": set, "TMO Manager": set})

    # Create TMO Area
    patches_and_areas = list[Patch]()
    patches_and_areas.append(build_tmo_area_object())

    # Create Patch objects from the data
    Config.logger.log(f"Creating {len(df)} Patch objects")
    for index, data_row in df.iterrows():
        new_patch = build_patch_object(data_row)

        # Check if the patch already exists
        existing_patch = next(
            (patch for patch in patches_and_areas if patch.name == new_patch.name), None)
        if existing_patch:
            # Merge responsible entities without duplicates
            existing_patch.responsibleEntities.extend(
                x for x in new_patch.responsibleEntities if x not in existing_patch.responsibleEntities)
        else:
            patches_and_areas.append(new_patch)

    # Export the patches to a JSON file
    patches_dicts = [asdict(patch) for patch in patches_and_areas]
    with open(Config.EXPORT_FILE_PATH, "w") as f:
        f.write(json.dumps(patches_dicts, indent=4))
    Config.logger.log(
        f"Exported patches to a file: ({Config.EXPORT_FILE_PATH})")

    # Save the patches to the database
    table = get_dynamodb_table(Config.TABLE_NAME, Config.STAGE)
    input(f"Saving {len(patches_and_areas)} patches and areas to the database.\n"
          f"Please confirm that the patches are correct in {Config.EXPORT_FILE_PATH}.\n"
          f"Hit enter to continue\n")
    for patch in patches_dicts:
        table.put_item(Item=patch)
    Config.logger.log(
        f"Patches saved to DynamoDB table {Config.TABLE_NAME} successfully")


def rollback_changes():
    table = get_dynamodb_table(Config.TABLE_NAME, Config.STAGE)
    with open(Config.EXPORT_FILE_PATH, "r") as f:
        patches = json.load(f)
    for patch in patches:
        table.delete_item(Key={"id": patch["id"]})
    Config.logger.log("Changes rolled back successfully.")


def main():
    create_fake_spreadsheet_data()
    create_tmo_patches()

    i = input("Would you like to rollback the changes? y/n\n")
    if (i.lower() == "y"):
        rollback_changes()
    Config.logger.log("Finished.")


if __name__ == "__main__":
    main()
